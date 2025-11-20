"""Utilities for parsing and cleaning bank statement PDFs using Adobe PDF Services API."""
from __future__ import annotations

import io
import os
import re
import sys
import tempfile
import traceback
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Dict, Iterable, List, Optional

import pandas as pd

try:
    import pdfplumber
except ImportError:
    pdfplumber = None  # type: ignore

from .adobe_pdf_service import AdobePDFService

# Функция для логирования в stderr (чтобы не мешать JSON в stdout)
def _log_debug(msg: str) -> None:
    """Выводит DEBUG лог в stderr."""
    print(msg, file=sys.stderr, flush=True)

@dataclass
class ProcessedTable:
    """Structured view of a filtered table extracted from the PDF."""

    page_number: int
    bank_name: Optional[str]
    rows: List[dict]


@dataclass
class StatementExtraction:
    """Combined result of metadata and filtered tables for a statement."""

    bank_name: Optional[str]
    metadata: Dict[str, str]
    tables: List[ProcessedTable]


class PDFStatementProcessor:
    """Extracts rows with non-empty credit column values from bank statements using Adobe PDF Services API."""

    def __init__(
        self,
        credit_headers: Optional[Iterable[str]] = None,
        debit_headers: Optional[Iterable[str]] = None,
        date_headers: Optional[Iterable[str]] = None,
        client_id: Optional[str] = None,
        client_secret: Optional[str] = None,
        credentials_file: Optional[str] = None,
        region: Optional[str] = None,
        connect_timeout: Optional[int] = None,
        read_timeout: Optional[int] = None,
    ) -> None:
        """
        Инициализация процессора с обязательным Adobe API.

        Args:
            credit_headers: Заголовки столбцов для кредита
            debit_headers: Заголовки столбцов для дебета
            date_headers: Заголовки столбцов для даты
            client_id: Adobe Client ID
            client_secret: Adobe Client Secret
            credentials_file: Путь к файлу credentials.json
            region: Регион обработки ('US' или 'EU')
            connect_timeout: Таймаут подключения в мс
            read_timeout: Таймаут чтения в мс
        """
        self._empty_tokens = {"", "-", "—", "none", "null", "nan", "н/д"}
        self.credit_headers = {
            self._normalize_header(h)
            for h in (credit_headers or {"credit", "кредит", "кредит сумма", "кредитование"})
        }
        self.debit_headers = {
            self._normalize_header(h)
            for h in (debit_headers or {"debit", "дебет", "дебет сумма"})
        }
        default_date_headers = {
            "дата",
            "дата операции",
            "дата документа",
            "date",
            "operation date",
            "transaction date",
            "data",
            "күні",
        }
        self.date_headers = {
            self._normalize_header(h)
            for h in (date_headers or default_date_headers)
        }

        # Инициализируем Adobe API сервис (обязательно)
        import sys
        print(f"[PDF_PROCESSOR] Инициализация AdobePDFService...", file=sys.stderr, flush=True)
        print(f"[PDF_PROCESSOR] client_id: {'✅ установлен' if client_id else '❌ не установлен'}", file=sys.stderr, flush=True)
        print(f"[PDF_PROCESSOR] client_secret: {'✅ установлен' if client_secret else '❌ не установлен'}", file=sys.stderr, flush=True)
        print(f"[PDF_PROCESSOR] credentials_file: {credentials_file or 'не установлен'}", file=sys.stderr, flush=True)
        print(f"[PDF_PROCESSOR] region: {region or 'US (по умолчанию)'}", file=sys.stderr, flush=True)
        
        try:
            self._adobe_service = AdobePDFService(
                client_id=client_id,
                client_secret=client_secret,
                credentials_file=credentials_file,
                region=region,
                connect_timeout=connect_timeout,
                read_timeout=read_timeout,
            )
            print(f"[PDF_PROCESSOR] ✅ AdobePDFService инициализирован успешно", file=sys.stderr, flush=True)
        except Exception as e:
            print(f"[PDF_PROCESSOR] ❌ Ошибка инициализации AdobePDFService: {e}", file=sys.stderr, flush=True)
            import traceback
            print(f"[PDF_PROCESSOR] Traceback: {traceback.format_exc()}", file=sys.stderr, flush=True)
            raise
        
        # Для сохранения последнего Excel файла для просмотра
        self._last_excel_bytes: Optional[bytes] = None
        self._last_excel_filename: Optional[str] = None

    @staticmethod
    def _normalize_header(header: str) -> str:
        return header.strip().lower().replace("\n", " ")

    @staticmethod
    def _is_row_empty(row: Iterable[object]) -> bool:
        return all((str(cell).strip() == "" or cell is None) for cell in row)

    def _is_effectively_empty(self, value: object) -> bool:
        if value is None:
            return True
        text = str(value).strip()
        return text.lower() in self._empty_tokens

    @staticmethod
    def _clean_numeric_value(value: str) -> str:
        """Очищает числовое значение от переносов строк и лишних пробелов."""
        # Убираем все переносы строк и заменяем на пустую строку
        cleaned = str(value).replace("\n", "").replace("\r", "")
        # Убираем неразрывные пробелы
        cleaned = cleaned.replace("\xa0", " ")
        # Убираем лишние пробелы между цифрами
        cleaned = re.sub(r"(\d)\s+(\d)", r"\1\2", cleaned)
        cleaned = cleaned.strip()
        
        # Проверяем на дублирование числа (например, "4150000,004150000,00" -> "4150000,00")
        # Ищем паттерн: число с запятой, затем то же число с запятой
        match = re.match(r"^(.+?,\d{2})\1$", cleaned)
        if match:
            cleaned = match.group(1)
            return cleaned
        
        # Также проверяем случай когда число без запятой дублируется
        # Например, "41500004150000" -> "4150000"
        match = re.match(r"^(\d+)\1$", cleaned.replace(",", "").replace(".", ""))
        if match:
            # Восстанавливаем формат с запятой, если был
            num_str = match.group(1)
            if "," in cleaned or "." in cleaned:
                # Пытаемся восстановить формат
                if len(num_str) >= 2:
                    num_str = num_str[:-2] + "," + num_str[-2:]
            cleaned = num_str
            return cleaned
        
        # Проверяем случай, когда два разных числа склеены вместе
        # Например, "33600000,0049563711,69" -> берем первое "33600000,00"
        # Ищем паттерн: число,запятая,две цифры, затем снова число,запятая,две цифры
        match = re.search(r"^(\d+,\d{2})(\d+,\d{2})$", cleaned)
        if match:
            # Берем первое число (можно изменить логику, если нужно брать последнее)
            cleaned = match.group(1)
            return cleaned
        
        # Также проверяем случай когда числа склеены без запятой между ними
        # Например, "33600000,0049563711,69" (но уже обработано выше) или "100000200000" (два числа)
        # Это более сложный случай - просто вернем исходное значение
        
        return cleaned

    @staticmethod
    def _to_decimal(value: str) -> Optional[Decimal]:
        cleaned = value.replace("\xa0", " ").strip()
        # Убираем переносы строк из числовых значений
        cleaned = cleaned.replace("\n", "").replace("\r", "")
        cleaned = re.sub(r"[^\d,.\-]", "", cleaned)
        cleaned = cleaned.replace(",", ".")
        if cleaned in {"", "-", "—", ".", "-.", ".-"}:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    def _detect_column(self, columns: Iterable[str], known_headers: Iterable[str]) -> Optional[int]:
        normalized_known = set(known_headers)
        for idx, col in enumerate(columns):
            if self._normalize_header(str(col)) in normalized_known:
                return idx
        for idx, col in enumerate(columns):
            normalized = self._normalize_header(str(col))
            if any(h in normalized for h in normalized_known):
                return idx
        return None

    def _extract_metadata(self, pdf) -> Dict[str, str]:
        """Извлечь метаданные из первой страницы PDF."""
        if pdfplumber is None or not hasattr(pdf, 'pages') or not pdf.pages:
            return {}

        first_page = pdf.pages[0]
        text = (first_page.extract_text() or "").replace("\xa0", " ")
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        header_lines: List[str] = []
        for line in lines:
            header_lines.append(line)
            if "номер" in line.lower() and "кредит" in line.lower():
                break
        metadata: Dict[str, str] = {}

        def assign(pattern: str, key: str) -> None:
            for line in header_lines:
                match = re.search(pattern, line, flags=re.IGNORECASE)
                if match:
                    metadata[key] = match.group(1).strip()
                    return

        assign(r"дата печати[:\s]+(.+)", "print_date")
        assign(r"время печати[:\s]+(.+)", "print_time")
        assign(r"клиент[:\s]+(.+)", "client")
        assign(r"бин/?иин[:\s]+(.+)", "bin_iin")
        assign(r"банк[:\s]+(.+)", "bank")
        assign(r"бик[:\s]+(.+)", "bic")
        assign(r"иик[:\s]+(.+)", "iik")
        assign(r"валюта[:\s]+(.+)", "currency")
        assign(r"входящий остаток[:\s]+(.+)", "opening_balance")
        assign(r"исходящий остаток[:\s]+(.+)", "closing_balance")

        combined_text = " ".join(header_lines)
        period_match = re.search(r"за период\s*с\s*([\d\.]+)\s*по\s*([\d\.]+)", combined_text, flags=re.IGNORECASE)
        if period_match:
            metadata["period_from"] = period_match.group(1)
            metadata["period_to"] = period_match.group(2)

        if "statement_title" not in metadata:
            for line in lines:
                if "выписка" in line.lower():
                    metadata["statement_title"] = line
                    break

        metadata["raw_header"] = header_lines
        return metadata

    def _looks_like_table_header(self, row: pd.Series, next_row: Optional[pd.Series] = None) -> bool:
        """
        Проверяет, выглядит ли строка как заголовок таблицы.
        
        Критерии заголовка таблицы:
        1. Содержит ключевое слово кредита/дебета
        2. Имеет достаточно непустых колонок (минимум 3-4)
        3. Содержит несколько ключевых слов таблицы (дата, номер, кредит/дебет)
        4. (Опционально) Следующая строка выглядит как данные
        """
        if row is None or row.empty:
            return False
        
        row_values = [str(cell).strip() for cell in row if pd.notna(cell)]
        non_empty_cells = [v for v in row_values if v and v.lower() not in self._empty_tokens]
        
        # Заголовок должен иметь минимум 3 непустых колонки
        if len(non_empty_cells) < 3:
            return False
        
        # Нормализуем все значения строки
        normalized_row = [self._normalize_header(cell) for cell in row_values]
        row_text = " ".join(normalized_row)
        
        # Проверяем наличие ключевых слов заголовков таблицы
        table_header_keywords = [
            "кредит", "дебет", "credit", "debit",
            "дата", "date", "күні",
            "номер", "документ", "document", "№",
            "назначение", "назнач",
        ]
        
        # Используем set для уникальности найденных ключевых слов
        found_keywords = {kw for kw in table_header_keywords if any(kw in cell for cell in normalized_row)}
        
        # Должно быть найдено минимум 2 различных ключевых слова (например, дата + кредит, или номер + кредит)
        if len(found_keywords) < 2:
            return False
        
        # Обязательно должно быть ключевое слово кредита или дебета
        has_credit_debit = any(
            any(header in cell for header in self.credit_headers | self.debit_headers)
            for cell in normalized_row
        )
        if not has_credit_debit:
            return False
        
        # Проверяем, что это НЕ метаданные
        metadata_keywords = [
            "лицевой счет", "лицевой счёт", "л/с",
            "валюта счета", "валюта", "currency",
            "период", "period",
            "входящий остаток", "исходящий остаток",
            "банк", "bank", "бик",
            "клиент", "client",
            "дата печати", "время печати",
            "выписка по счету", "выписка",
        ]
        
        # Если строка содержит только метаданные без достаточного количества колонок таблицы
        contains_metadata_only = any(
            any(mk in cell for cell in normalized_row) for mk in metadata_keywords
        ) and len(found_keywords) < 3
        
        if contains_metadata_only:
            return False
        
        # Дополнительная проверка: если следующая строка предоставлена, она должна выглядеть как данные
        if next_row is not None and not next_row.empty:
            next_row_values = [str(cell).strip() for cell in next_row if pd.notna(cell)]
            next_row_text = " ".join(next_row_values).lower()
            
            # Следующая строка должна содержать признаки данных: цифры, даты, но не заголовки
            has_numbers = any(re.search(r'\d', val) for val in next_row_values)
            has_no_header_keywords = not any(
                kw in " ".join(next_row_values).lower() 
                for kw in ["кредит", "дебет", "дата", "номер", "credit", "debit", "date"]
            )
            
            if has_numbers and has_no_header_keywords:
                return True  # Следующая строка похожа на данные
        
        # Если следующей строки нет или она не похожа на данные, все равно это может быть заголовок
        # если выполнены все основные критерии (проверены выше):
        # - минимум 3 непустых колонки
        # - минимум 2 ключевых слова
        # - есть кредит/дебет
        # - это не метаданные
        # Если выполнены все эти условия, это похоже на заголовок
        # Более строгая проверка: если много колонок (>=4) или много ключевых слов (>=3), это точно заголовок
        # Если меньше - это все равно может быть заголовок, если выполнены все предыдущие критерии
        # (т.к. все критерии уже проверены выше, просто возвращаем True)
        return True

    def _find_header_row(
        self, dataframe: pd.DataFrame
    ) -> tuple[Optional[int], Optional[pd.Series], bool]:
        if dataframe.empty:
            return None, None, False

        max_rows_to_check = min(len(dataframe), 50)
        
        # ШАГ 1: Прямой поиск заголовка - ищем строку, которая явно похожа на заголовок таблицы
        # Проверяем каждую строку отдельно, без накопления
        for idx in range(max_rows_to_check):
            row = dataframe.iloc[idx].fillna("").astype(str)
            
            # Берем следующую строку для дополнительной проверки (если есть)
            next_row = None
            if idx + 1 < len(dataframe):
                next_row = dataframe.iloc[idx + 1].fillna("").astype(str)
            
            # Проверяем, выглядит ли эта строка как заголовок таблицы
            if self._looks_like_table_header(row, next_row):
                _log_debug(f"[DEBUG] Найден заголовок прямой проверкой (строка {idx}): выглядит как заголовок таблицы")
                # Дополнительно проверяем наличие кредита/дебета для уверенности
                normalized = [self._normalize_header(str(cell)) for cell in row]
                has_credit = any(
                    any(header in cell for header in self.credit_headers | self.debit_headers)
                    for cell in normalized
                )
                if has_credit:
                    return idx, row, True

        # ШАГ 2: Fallback - используем накопление только если прямой поиск не дал результата
        # НО: при накоплении тоже проверяем, что результат выглядит как заголовок таблицы
        _log_debug(f"[DEBUG] Прямой поиск не дал результата, пробуем накопление с проверкой")
        
        accumulated: Optional[pd.Series] = None
        accumulated_start_idx = None
        
        for idx in range(min(len(dataframe), 20)):  # Ограничиваем накопление 20 строками
            row = dataframe.iloc[idx].fillna("").astype(str)
            
            if accumulated is None:
                accumulated = row.copy()
                accumulated_start_idx = idx
            else:
                accumulated = accumulated.combine(
                    row,
                    lambda a, b: " ".join(filter(None, [str(a).strip(), str(b).strip()])),
                    fill_value="",
                )
            
            # Проверяем накопленную строку - она должна выглядеть как заголовок таблицы
            next_row = None
            if idx + 1 < len(dataframe):
                next_row = dataframe.iloc[idx + 1].fillna("").astype(str)
            
            if self._looks_like_table_header(accumulated, next_row):
                _log_debug(f"[DEBUG] Найден заголовок накоплением (строки {accumulated_start_idx}-{idx}): выглядит как заголовок таблицы")
                return idx, accumulated, True

        # ШАГ 3: Последний fallback - ищем первую строку с кредитом/дебетом (без строгой проверки)
        _log_debug(f"[DEBUG] Накопление не дало результата, ищем первую строку с кредитом/дебетом")
        
        for idx in range(min(len(dataframe), 50)):
            row = dataframe.iloc[idx].fillna("").astype(str)
            normalized = [self._normalize_header(str(cell)) for cell in row]
            
            # Ищем просто наличие кредита/дебета
            if any(cell in self.credit_headers | self.debit_headers for cell in normalized):
                _log_debug(f"[DEBUG] Найдена строка {idx} с кредитом/дебетом как fallback")
                return idx, row, True
            if any(
                any(header in cell for header in self.credit_headers | self.debit_headers)
                for cell in normalized
            ):
                _log_debug(f"[DEBUG] Найдена строка {idx} с кредитом/дебетом (частичное совпадение) как fallback")
                return idx, row, True

        # Последний fallback - первая строка
        fallback = dataframe.iloc[0].fillna("").astype(str)
        _log_debug(f"[DEBUG] Заголовок не найден, используем первую строку как последний fallback")
        return None, fallback, False

    def _prepare_columns(self, header_series: pd.Series) -> List[str]:
        cleaned_columns: List[str] = []
        seen: set[str] = set()
        for raw_value in header_series:
            col = str(raw_value).replace("\n", " ").strip()
            normalized = col.lower()
            normalized = normalized.replace("ё", "е")
            if normalized in self._empty_tokens or col == "":
                cleaned_columns.append("")
                continue
            if "кредит" in normalized:
                col = "Кредит"
                normalized = col.lower()
            elif "дебет" in normalized:
                col = "Дебет"
                normalized = col.lower()
            elif any(token in normalized for token in ("назнач", "тағайындал", "төлем")):
                col = "Назначение платежа"
                normalized = col.lower()
            # Сначала проверяем дату, чтобы не путать её с курсом
            elif any(token in normalized for token in ("дата", "күні", "date")):
                col = "Дата"
                normalized = col.lower()
            # Курс проверяем после даты, но только если это не дата
            elif any(token in normalized for token in ("курс", "бағам")) and "дата" not in normalized:
                col = "Курс"
                normalized = col.lower()
            elif any(token in normalized for token in ("отправ", "жібер")):
                col = "Отправитель"
                normalized = col.lower()
            elif any(token in normalized for token in ("получ", "алушы")):
                col = "Получатель"
                normalized = col.lower()
            elif any(token in normalized for token in ("номер", "нөмір", "құжат", "document")):
                col = "Документ"
                normalized = col.lower()
            elif "№" in col or any(token in normalized for token in ("номер", "no", "entry")):
                col = "№"
                normalized = col.lower()
            if normalized in seen:
                suffix = 2
                candidate = f"{col}_{suffix}"
                while candidate.lower() in seen:
                    suffix += 1
                    candidate = f"{col}_{suffix}"
                col = candidate
                normalized = col.lower()
            cleaned_columns.append(col)
            seen.add(normalized)
        return cleaned_columns

    @staticmethod
    def _is_numeric_header(header: str) -> bool:
        compact = re.sub(r"[^\d]", "", header)
        return bool(compact) and compact == re.sub(r"[^\d]", "", header)

    def _process_dataframe_with_repeated_headers(
        self,
        dataframe: pd.DataFrame,
        page_number: int,
        bank_name: Optional[str],
    ) -> List[Optional[ProcessedTable]]:
        """
        Обрабатывает DataFrame, разделяя его на части по повторяющимся заголовкам.
        Возвращает список обработанных таблиц (одну для каждой секции с заголовком).
        """
        if dataframe.empty:
            return []
        
        results: List[Optional[ProcessedTable]] = []
        fallback_columns: Optional[List[str]] = None
        
        # Ищем все строки, которые выглядят как заголовки во ВСЕМ DataFrame
        # Это важно для длинных выписок, где заголовки повторяются на каждой странице
        header_indices = []
        
        # Проверяем все строки DataFrame для поиска повторяющихся заголовков
        for idx in range(len(dataframe)):
            row = dataframe.iloc[idx].fillna("").astype(str)
            if idx + 1 < len(dataframe):
                next_row = dataframe.iloc[idx + 1].fillna("").astype(str)
            else:
                next_row = None
            
            if self._looks_like_table_header(row, next_row):
                # Проверяем, что это не просто случайное совпадение
                normalized = [self._normalize_header(str(cell)) for cell in row]
                has_credit = any(
                    any(header in cell for header in self.credit_headers | self.debit_headers)
                    for cell in normalized
                )
                if has_credit:
                    # Проверяем, что это не дубликат предыдущего заголовка (похожая структура)
                    # Если предыдущий заголовок был недавно (в пределах 5 строк), это может быть дубликат
                    is_duplicate = False
                    if header_indices:
                        last_header_idx = header_indices[-1]
                        if idx - last_header_idx < 5:
                            # Сравниваем содержимое заголовков
                            last_header_row = dataframe.iloc[last_header_idx].fillna("").astype(str)
                            last_normalized = [self._normalize_header(str(cell)) for cell in last_header_row]
                            # Если заголовки очень похожи - это дубликат
                            if len(set(normalized) & set(last_normalized)) >= 3:
                                is_duplicate = True
                    
                    if not is_duplicate:
                        header_indices.append(idx)
        
        _log_debug(f"[DEBUG] Найдено заголовков в листе: {len(header_indices)} (индексы: {header_indices[:20]}...)" if len(header_indices) > 20 else f"[DEBUG] Найдено заголовков в листе: {len(header_indices)} (индексы: {header_indices})")
        
        # Если найдено несколько заголовков, разбиваем DataFrame на части
        # Каждая секция будет обработана с правильными заголовками
        if len(header_indices) > 1:
            _log_debug(f"[DEBUG] Найдено {len(header_indices)} заголовков, разбиваю DataFrame на {len(header_indices)} секций")
            for i, header_idx in enumerate(header_indices):
                start_idx = header_idx
                # Берем все строки до следующего заголовка или до конца
                if i + 1 < len(header_indices):
                    end_idx = header_indices[i + 1]
                else:
                    end_idx = len(dataframe)
                
                section_df = dataframe.iloc[start_idx:end_idx].copy().reset_index(drop=True)
                _log_debug(f"[DEBUG] Обрабатываю секцию {i+1}/{len(header_indices)}: строки {start_idx}-{end_idx} (размер секции: {len(section_df)} строк)")
                
                processed, fallback_columns = self._process_dataframe(
                    section_df, 
                    page_number=page_number, 
                    bank_name=bank_name, 
                    fallback_columns=fallback_columns
                )
                
                if processed:
                    results.append(processed)
                    _log_debug(f"[DEBUG] Секция {i+1} обработана: найдено {len(processed.rows)} строк с кредитом")
        else:
            # Заголовок один или не найден - обрабатываем весь DataFrame
            # Это нормально для банков, где заголовок только в начале выписки
            _log_debug(f"[DEBUG] Заголовок один ({len(header_indices)}) или не найден - обрабатываю весь DataFrame как одну секцию")
            processed, _ = self._process_dataframe(
                dataframe, 
                page_number=page_number, 
                bank_name=bank_name, 
                fallback_columns=fallback_columns
            )
            if processed:
                results.append(processed)
        
        return results
    
    def _process_dataframe(
        self,
        dataframe: pd.DataFrame,
        page_number: int,
        bank_name: Optional[str],
        fallback_columns: Optional[List[str]] = None,
    ) -> tuple[Optional[ProcessedTable], Optional[List[str]]]:
        if dataframe.empty:
            _log_debug(f"[DEBUG] DataFrame пустой")
            return None, fallback_columns

        _log_debug(f"[DEBUG] Ищу заголовок в DataFrame: {len(dataframe)} строк, {len(dataframe.columns)} колонок")
        header_idx, header_series, header_found = self._find_header_row(dataframe)
        _log_debug(f"[DEBUG] Найден заголовок: idx={header_idx}, found={header_found}")

        if header_found and header_series is not None and header_idx is not None:
            dataframe = dataframe.iloc[header_idx + 1 :].reset_index(drop=True)
            columns = self._prepare_columns(header_series)
        elif header_series is not None and not fallback_columns:
            columns = self._prepare_columns(header_series)
            drop_from = (header_idx + 1) if header_idx is not None else 1
            if drop_from > 0:
                dataframe = dataframe.iloc[drop_from:].reset_index(drop=True)
        elif fallback_columns:
            columns = fallback_columns
        else:
            return None, fallback_columns

        if not columns:
            return None, fallback_columns

        col_count = len(columns)
        if col_count < dataframe.shape[1]:
            extras = [f"extra_{i}" for i in range(dataframe.shape[1] - col_count)]
            columns = columns + extras
        elif col_count > dataframe.shape[1]:
            columns = columns[: dataframe.shape[1]]

        dataframe.columns = columns
        dataframe = dataframe[[col for col in dataframe.columns if col]]

        dataframe = dataframe.replace(pd.NA, None)
        dataframe = self._consolidate_rows(dataframe)
        if dataframe.empty:
            _log_debug(f"[DEBUG] DataFrame пустой после консолидации")
            return None, fallback_columns

        _log_debug(f"[DEBUG] Ищу колонку кредита среди колонок: {list(dataframe.columns)}")
        _log_debug(f"[DEBUG] Ищу колонку кредита среди заголовков: {self.credit_headers}")
        credit_column_idx = self._detect_column(dataframe.columns, self.credit_headers)
        if credit_column_idx is None:
            print(f"[ERROR] Не найдена колонка кредита! Доступные колонки: {list(dataframe.columns)}", file=sys.stderr, flush=True)
            return None, fallback_columns
        
        _log_debug(f"[DEBUG] Найдена колонка кредита: idx={credit_column_idx}, name={dataframe.columns[credit_column_idx]}")

        date_column_idx = self._detect_column(dataframe.columns, self.date_headers)
        debit_column_idx = self._detect_column(dataframe.columns, self.debit_headers)

        credit_column_name = dataframe.columns[credit_column_idx]

        filtered_rows: List[dict] = []

        # Ключевые слова для фильтрации итоговых строк
        summary_keywords = [
            "обороты", "итого", "входящий остаток", "исходящий остаток",
            "всего", "total", "summary", "итог", "остаток",
            "документов по дебету", "документов по кредиту", "документов:"
        ]

        row_number = 0
        for idx, row in dataframe.iterrows():
            row_number += 1
            if self._is_row_empty(row):
                _log_debug(f"[DEBUG] Строка {row_number}: пропущена - пустая строка")
                continue

            # Сначала проверяем кредит - если есть кредит > 0, это может быть реальная операция
            credit_cell = row.iloc[credit_column_idx]
            has_credit = False
            credit_value = None
            amount = None
            
            if pd.notna(credit_cell) and not self._is_effectively_empty(credit_cell):
                credit_value = self._clean_numeric_value(credit_cell)
                if credit_value:
                    amount = self._to_decimal(credit_value)
                    if amount is not None and amount > 0:
                        has_credit = True
            
            # Получаем номер документа и дату для проверки
            doc_no = None
            date_val = None
            has_doc_no = False
            has_date = False
            
            if "№" in row.index:
                doc_no_cell = row["№"]
                if pd.notna(doc_no_cell) and not self._is_effectively_empty(doc_no_cell):
                    doc_no = str(doc_no_cell).strip()
                    # Проверяем, что номер документа содержит цифры (это реальный номер, а не заголовок)
                    if doc_no and re.search(r"\d", doc_no):
                        has_doc_no = True
            
            if date_column_idx is not None:
                date_cell = row.iloc[date_column_idx]
                if pd.notna(date_cell) and not self._is_effectively_empty(date_cell):
                    date_val = str(date_cell).strip()[:20]
                    # Проверяем, что дата содержит паттерн даты (например, 2024-05-06 или 06.05.2024)
                    if date_val and (re.search(r"\d{4}[-/]\d{2}[-/]\d{2}", date_val) or re.search(r"\d{2}\.\d{2}\.\d{4}", date_val)):
                        has_date = True
            
            # Проверяем, не является ли это итоговой строкой
            # НО: если есть кредит > 0 и номер документа - это реальная операция, не итоговая (даже если дата не валидна)
            row_text = " ".join([str(val) for val in row.values if pd.notna(val)]).lower()
            contains_summary_keywords = any(keyword in row_text for keyword in summary_keywords)
            
            # Если это выглядит как реальная операция (есть кредит и номер документа), не пропускаем
            # Дата может быть не валидна из-за переносов строк, но это не значит, что это итоговая строка
            if contains_summary_keywords and not (has_credit and has_doc_no):
                _log_debug(f"[DEBUG] Строка {row_number}: пропущена - итоговая строка (содержит: {[kw for kw in summary_keywords if kw in row_text]}, кредит: {has_credit}, №: {has_doc_no}, дата: {has_date})")
                continue

            # Если нет кредита, пропускаем
            if not has_credit:
                _log_debug(f"[DEBUG] Строка {row_number}: пропущена - нет кредита (№ документа: {doc_no})")
                continue
            
            _log_debug(f"[DEBUG] Строка {row_number}: найдена с кредитом {amount} (№: {doc_no}, Дата: {date_val if date_val else 'нет'}, значение: {credit_value})")

            # Если есть кредит и номер документа - это реальная операция, не пропускаем из-за даты
            # Дата может быть в любом формате или вообще отсутствовать - это не проблема
            if date_column_idx is not None:
                date_cell = row.iloc[date_column_idx]
                if pd.isna(date_cell) or self._is_effectively_empty(date_cell):
                    # Если нет даты, но есть кредит и номер документа - оставляем строку
                    if has_credit and has_doc_no:
                        _log_debug(f"[DEBUG] Строка {row_number}: дата пустая, но есть кредит и № документа - оставляем (№: {doc_no}, кредит: {amount})")
                    else:
                        _log_debug(f"[DEBUG] Строка {row_number}: пропущена - нет даты и нет кредита/№ документа (№: {doc_no})")
                        continue
                else:
                    # Извлекаем дату из ячейки, игнорируя текст после итоговых слов
                    date_str_full = str(date_cell).strip()
                    date_str_lower = date_str_full.lower()
                    
                    # Ищем позицию первого вхождения итоговых слов
                    summary_positions = []
                    for keyword in ["обороты", "итого", "документов"]:
                        pos = date_str_lower.find(keyword)
                        if pos >= 0:
                            summary_positions.append(pos)
                    
                    # Если нашли итоговые слова, берем только часть до них
                    if summary_positions:
                        min_pos = min(summary_positions)
                        date_str_clean = date_str_full[:min_pos].strip()
                        _log_debug(f"[DEBUG] Строка {row_number}: дата содержит итоговые слова, извлечена дата: '{date_str_clean}' из '{date_str_full[:50]}...'")
                    else:
                        date_str_clean = date_str_full
                    
                    # Проверяем, что извлеченная дата валидна (содержит паттерн даты)
                    if not (re.search(r"\d{4}[-/]\d{2}[-/]\d{2}", date_str_clean) or re.search(r"\d{2}\.\d{2}\.\d{4}", date_str_clean)):
                        # Если дата не валидна, но есть кредит и номер документа - оставляем строку
                        # Дата может быть в странном формате (например, "YYYY-00-DD 00:00:SS"), это не проблема
                        if has_credit and has_doc_no:
                            _log_debug(f"[DEBUG] Строка {row_number}: дата в странном формате '{date_str_clean}', но есть кредит и № документа - оставляем (№: {doc_no}, кредит: {amount})")
                        else:
                            _log_debug(f"[DEBUG] Строка {row_number}: пропущена - дата не валидна и нет кредита/№ документа (№: {doc_no}, дата: '{date_str_clean}')")
                            continue

            # Проверяем дебет - пропускаем строку если есть дебет
            # Если есть и кредит, и дебет → пропускаем (это может быть дубликат или ошибка парсинга)
            # Если есть только кредит (без дебета) → включаем в результат
            if debit_column_idx is not None:
                debit_cell = row.iloc[debit_column_idx]
                if pd.notna(debit_cell) and not self._is_effectively_empty(debit_cell):
                    debit_value = self._clean_numeric_value(debit_cell)
                    debit_amount = self._to_decimal(debit_value) if debit_value else None
                    # Если есть реальный дебет > 0
                    if debit_amount is not None and debit_amount > 0:
                        # Если есть и кредит, и дебет - пропускаем
                        if amount is not None and amount > 0:
                            _log_debug(f"[DEBUG] Строка {row_number}: пропущена - есть кредит {amount} и дебет {debit_amount} (№: {doc_no})")
                            continue
                        # Если только дебет, нет кредита - пропускаем
                        _log_debug(f"[DEBUG] Строка {row_number}: пропущена - есть дебет {debit_amount}, но нет кредита (№: {doc_no})")
                        continue
                    elif debit_amount is None and debit_value.lower() not in self._empty_tokens:
                        # Дебет не распознан как число, но есть текст
                        # Пропускаем только если нет кредита
                        if amount is None or amount == 0:
                            _log_debug(f"[DEBUG] Строка {row_number}: пропущена - дебет не распознан, нет кредита (№: {doc_no})")
                            continue

            sanitized_row: Dict[str, str] = {}
            for key, value in row.items():
                column_name = str(key).strip()
                if column_name == "":
                    continue
                # Пропускаем числовые заголовки, но ОСТАВЛЯЕМ "Дебет", "Кредит", "Курс"
                if (column_name != credit_column_name and 
                    self._is_numeric_header(column_name) and
                    column_name not in ("Дебет", "Кредит", "Курс")):
                    continue
                if pd.isna(value) or self._is_effectively_empty(value):
                    continue
                
                # Для числовых колонок (Кредит, Дебет, Курс) убираем переносы строк и дублирование
                if column_name in ("Дебет", "Кредит", "Курс"):
                    sanitized_row[column_name] = self._clean_numeric_value(value)
                # Для колонки "№" убираем переносы строк, но оставляем пробелы
                elif column_name == "№":
                    cleaned = str(value).replace("\n", "").replace("\r", "").strip()
                    sanitized_row[column_name] = cleaned
                # Для колонки "Дата" убираем переносы строк, заменяя на пробел
                elif column_name == "Дата":
                    cleaned = str(value).replace("\n", " ").replace("\r", " ").strip()
                    # Убираем лишние пробелы
                    cleaned = re.sub(r"\s+", " ", cleaned)
                    # Убираем текст после итоговых слов (если есть)
                    date_str_lower = cleaned.lower()
                    summary_positions = []
                    for keyword in ["обороты", "итого", "документов"]:
                        pos = date_str_lower.find(keyword)
                        if pos >= 0:
                            summary_positions.append(pos)
                    if summary_positions:
                        min_pos = min(summary_positions)
                        cleaned = cleaned[:min_pos].strip()
                    sanitized_row[column_name] = cleaned
                else:
                    # Для остальных колонок оставляем как есть (могут быть переносы в тексте)
                    sanitized_row[column_name] = str(value).strip()

            # credit_value уже очищен через _clean_numeric_value
            sanitized_row[credit_column_name] = credit_value
            
            if date_column_idx is not None:
                date_column_name = dataframe.columns[date_column_idx]
                date_value = str(row.iloc[date_column_idx]).strip()
                # Очищаем дату от переносов строк
                date_value = date_value.replace("\n", " ").replace("\r", " ").strip()
                date_value = re.sub(r"\s+", " ", date_value)
                # Убираем текст после итоговых слов (если есть)
                date_str_lower = date_value.lower()
                summary_positions = []
                for keyword in ["обороты", "итого", "документов"]:
                    pos = date_str_lower.find(keyword)
                    if pos >= 0:
                        summary_positions.append(pos)
                if summary_positions:
                    min_pos = min(summary_positions)
                    date_value = date_value[:min_pos].strip()
                sanitized_row[date_column_name] = date_value

            if not sanitized_row:
                _log_debug(f"[DEBUG] Строка {row_number}: пропущена - пустой sanitized_row после обработки (№: {doc_no})")
                continue

            _log_debug(f"[DEBUG] Строка {row_number}: ✅ ДОБАВЛЕНА в результат (№: {doc_no}, Кредит: {amount})")
            filtered_rows.append(sanitized_row)

        if not filtered_rows:
            _log_debug(f"[DEBUG] Не найдено строк для обработки")
            return None, fallback_columns

        _log_debug(f"[DEBUG] Итого добавлено в результат: {len(filtered_rows)} строк")
        
        # Выводим номера документов из результата для проверки
        doc_numbers = []
        for row in filtered_rows:
            if "№" in row:
                doc_numbers.append(row["№"])
        _log_debug(f"[DEBUG] Номера документов в результате: {doc_numbers[:10]}..." if len(doc_numbers) > 10 else f"[DEBUG] Номера документов в результате: {doc_numbers}")

        return (
            ProcessedTable(page_number=page_number, bank_name=bank_name, rows=filtered_rows),
            dataframe.columns.tolist(),
        )

    def _consolidate_rows(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        if dataframe.empty:
            return dataframe

        _log_debug(f"[DEBUG] Консолидация строк: было {len(dataframe)} строк")
        columns = list(dataframe.columns)
        date_pattern = re.compile(r"\d{2}\.\d{2}\.\d{2,4}")
        results: List[Dict[str, Optional[str]]] = []
        current: Optional[Dict[str, Optional[str]]] = None
        skipped_empty = 0
        skipped_headers = 0
        processed_rows = 0

        for row in dataframe.itertuples(index=False, name=None):
            values = {}
            for col, val in zip(columns, row):
                if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
                    values[col] = None
                else:
                    values[col] = str(val).strip()

            if all(value is None for value in values.values()):
                skipped_empty += 1
                continue

            row_text = " ".join(
                value.lower() for value in values.values() if value
            )
            if (
                any(token in row_text for token in ("дата", "назна", "дебет", "кредит"))
                and not any(char.isdigit() for char in row_text)
            ):
                skipped_headers += 1
                continue
            
            processed_rows += 1

            candidate_no = values.get("№") or values.get("Номер док") or values.get("Документ")
            candidate_date = values.get("Дата")
            candidate_credit = values.get("Кредит")
            candidate_debit = values.get("Дебет")
            
            # Проверяем, есть ли кредит или дебет в строке
            has_credit_value = candidate_credit and candidate_credit.strip() and candidate_credit.strip() not in ("0", "0,00", "0.00", "")
            has_debit_value = candidate_debit and candidate_debit.strip() and candidate_debit.strip() not in ("0", "0,00", "0.00", "")
            
            new_entry = False
            extracted_no: Optional[str] = None

            if candidate_no and re.search(r"\d", candidate_no):
                new_entry = True
                extracted_no = candidate_no
                if values.get("№"):
                    values["№"] = None
            elif candidate_date:
                match = re.match(r"(\d+)\s+(.*)", candidate_date)
                if match:
                    extracted_no = match.group(1)
                    values["Дата"] = match.group(2)
                    new_entry = True
                elif date_pattern.search(candidate_date):
                    new_entry = True
            # Если есть кредит или дебет - это тоже может быть новая запись
            elif has_credit_value or has_debit_value:
                new_entry = True
                # Пробуем найти номер документа в других колонках
                for col in ["Номер док", "Документ", "№"]:
                    if values.get(col) and re.search(r"\d", str(values.get(col))):
                        extracted_no = str(values.get(col))
                        break

            if new_entry:
                if current:
                    results.append(current)
                current = {col: None for col in columns}
                if extracted_no:
                    current["№"] = extracted_no
                if values.get("Дата"):
                    current["Дата"] = values["Дата"]
                for col in columns:
                    if col in ("№", "Дата"):
                        continue
                    value = values.get(col)
                    if value:
                        if col in ("Дебет", "Кредит", "Курс"):
                            current[col] = value.replace(" ", "")
                        else:
                            current[col] = value
                continue

            if not current:
                continue

            for col, value in values.items():
                if not value or col == "№":
                    continue
                if col == "Дата":
                    combined = " ".join(filter(None, [current.get(col), value])).strip()
                    current[col] = combined
                elif col in ("Дебет", "Кредит", "Курс"):
                    combined = (current.get(col) or "") + value.replace(" ", "")
                    current[col] = combined
                else:
                    combined = " ".join(filter(None, [current.get(col), value])).strip()
                    current[col] = combined

        if current:
            results.append(current)

        _log_debug(f"[DEBUG] Консолидация завершена: обработано {processed_rows} строк, пропущено пустых {skipped_empty}, пропущено заголовков {skipped_headers}, создано записей {len(results)}")
        
        result_df = pd.DataFrame(results)
        result_df = result_df.replace("", pd.NA).dropna(how="all")
        _log_debug(f"[DEBUG] После dropna: осталось {len(result_df)} строк")
        return result_df.reset_index(drop=True)

    def extract(self, pdf_bytes: bytes, bank_name: Optional[str] = None) -> StatementExtraction:
        """
        Извлечь данные из PDF через Adobe API: PDF → XLSX → обработка → результат.

        Процесс:
        1. PDF файл отправляется в Adobe API
        2. Adobe API конвертирует PDF в Excel (XLSX)
        3. Excel файл обрабатывается для извлечения строк с кредитом
        4. Возвращается структурированный результат

        Args:
            pdf_bytes: Байты PDF файла
            bank_name: Имя банка (опционально)

        Returns:
            StatementExtraction с извлеченными данными
        """
        metadata: Dict[str, str] = {}
        tables: List[ProcessedTable] = []

        try:
            # ШАГ 1: Конвертируем PDF в Excel через Adobe API
            import sys
            print(f"[PDF_PROCESSOR] ========== НАЧАЛО ИЗВЛЕЧЕНИЯ ==========", file=sys.stderr, flush=True)
            print(f"[PDF_PROCESSOR] Размер PDF: {len(pdf_bytes)} байт", file=sys.stderr, flush=True)
            print(f"[PDF_PROCESSOR] Имя файла: {bank_name or 'не указано'}", file=sys.stderr, flush=True)
            print(f"[PDF_PROCESSOR] Отправка PDF в Adobe API для конвертации в Excel...", file=sys.stderr, flush=True)
            
            # Получаем исходные байты Excel файла
            excel_bytes = self._adobe_service.convert_pdf_to_excel(pdf_bytes, filename=bank_name)
            print(f"[PDF_PROCESSOR] ✅ Excel файл получен от Adobe API: {len(excel_bytes)} байт", file=sys.stderr, flush=True)
            
            # Сохраняем исходный Excel файл для возможности просмотра
            self._save_last_excel_bytes(excel_bytes, bank_name)
            
            # Конвертируем Excel байты в DataFrame для обработки
            print(f"[PDF_PROCESSOR] Чтение Excel файла в DataFrame...", file=sys.stderr, flush=True)
            excel_file = io.BytesIO(excel_bytes)
            
            # Читаем все листы из Excel файла
            from openpyxl import load_workbook
            wb = load_workbook(excel_file, data_only=True)
            sheet_names = wb.sheetnames
            print(f"[PDF_PROCESSOR] Найдено листов в Excel: {len(sheet_names)}", file=sys.stderr, flush=True)
            excel_file.seek(0)  # Сбрасываем позицию для чтения через pandas

            # ШАГ 2: Извлекаем метаданные из PDF (опционально, если pdfplumber доступен)
            if pdfplumber is not None:
                try:
                    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                        metadata = self._extract_metadata(pdf)
                except Exception as e:
                    import sys
                    print(f"[WARNING] Не удалось извлечь метаданные: {e}", file=sys.stderr, flush=True)
                    metadata = {}
            else:
                metadata = {}

            metadata.setdefault("bank_name", bank_name or "")
            metadata["extraction_method"] = "adobe_pdf_services_api"

            # ШАГ 3: Обрабатываем каждый лист Excel отдельно
            for sheet_idx, sheet_name in enumerate(sheet_names):
                excel_file.seek(0)  # Сбрасываем позицию для каждого листа
                try:
                    excel_df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
                    print(f"[PDF_PROCESSOR] ✅ Лист '{sheet_name}' прочитан: {len(excel_df)} строк, {len(excel_df.columns)} колонок", file=sys.stderr, flush=True)
                    
                    if excel_df.empty:
                        print(f"[PDF_PROCESSOR] Лист '{sheet_name}' пустой, пропускаем", file=sys.stderr, flush=True)
                        continue
                    
                    _log_debug(f"[DEBUG] Начинаю обработку листа '{sheet_name}': {len(excel_df)} строк, {len(excel_df.columns)} колонок")
                    _log_debug(f"[DEBUG] Колонки: {list(excel_df.columns)}")
                    
                    # Обрабатываем лист - ищем все повторяющиеся заголовки и разбиваем на секции
                    # Это важно для выписок, где на каждой странице PDF есть заголовки столбцов
                    processed_tables = self._process_dataframe_with_repeated_headers(
                        excel_df, 
                        page_number=sheet_idx + 1, 
                        bank_name=bank_name
                    )
                    
                    for processed in processed_tables:
                        if processed:
                            tables.append(processed)
                            print(f"[INFO] Извлечено {len(processed.rows)} строк с кредитом с листа '{sheet_name}'", file=sys.stderr, flush=True)
                        else:
                            print(f"[WARNING] Не удалось обработать часть листа '{sheet_name}': processed вернул None", file=sys.stderr, flush=True)
                            
                except Exception as e:
                    print(f"[ERROR] Ошибка при обработке листа '{sheet_name}': {e}", file=sys.stderr, flush=True)
                    print(f"[ERROR] Traceback: {traceback.format_exc()}", file=sys.stderr, flush=True)
                    # Продолжаем обработку остальных листов даже если один упал
                    continue

        except Exception as e:
            import sys
            print(f"[PDF_PROCESSOR] ❌ Ошибка при обработке через Adobe API: {e}", file=sys.stderr, flush=True)
            import traceback
            print(f"[PDF_PROCESSOR] Traceback: {traceback.format_exc()}", file=sys.stderr, flush=True)
            metadata["extraction_method"] = "adobe_pdf_services_api_failed"
            metadata["error"] = str(e)
            raise  # Пробрасываем ошибку, т.к. у нас нет fallback

        import sys
        print(f"[PDF_PROCESSOR] ========== ЗАВЕРШЕНИЕ ИЗВЛЕЧЕНИЯ ==========", file=sys.stderr, flush=True)
        print(f"[PDF_PROCESSOR] Найдено таблиц: {len(tables)}", file=sys.stderr, flush=True)
        total_rows = sum(len(table.rows) for table in tables)
        print(f"[PDF_PROCESSOR] Всего строк с кредитом: {total_rows}", file=sys.stderr, flush=True)
        print(f"[PDF_PROCESSOR] Метаданные: {len(metadata)} ключей", file=sys.stderr, flush=True)
        
        return StatementExtraction(bank_name=bank_name, metadata=metadata, tables=tables)

    def _save_last_excel_bytes(self, excel_bytes: bytes, filename: Optional[str] = None) -> None:
        """Сохраняет последний Excel файл (исходные байты) в память для возможности просмотра."""
        try:
            self._last_excel_bytes = excel_bytes
            # Формируем имя файла с расширением .xlsx
            if filename:
                if not filename.endswith('.xlsx'):
                    if filename.endswith('.pdf'):
                        filename = filename.replace('.pdf', '.xlsx')
                    else:
                        filename = f"{filename}.xlsx"
                self._last_excel_filename = filename
            else:
                self._last_excel_filename = "converted.xlsx"
            import sys
            print(f"[INFO] Excel файл сохранен для просмотра: {self._last_excel_filename} (размер: {len(excel_bytes)} байт)", file=sys.stderr, flush=True)
        except Exception as e:
            import sys
            print(f"[WARNING] Не удалось сохранить Excel файл для просмотра: {e}", file=sys.stderr, flush=True)
    
    def get_last_excel(self) -> tuple[Optional[bytes], Optional[str]]:
        """Возвращает последний Excel файл и его имя."""
        return self._last_excel_bytes, self._last_excel_filename


def merge_tables(tables: Iterable[ProcessedTable]) -> pd.DataFrame:
    """Merge processed tables into a single dataframe."""
    normalized_rows = []
    for table in tables:
        for row in table.rows:
            enriched_row = {
                "page_number": table.page_number,
                "bank_name": table.bank_name,
                **row,
            }
            normalized_rows.append(enriched_row)
    if not normalized_rows:
        return pd.DataFrame()
    return pd.DataFrame(normalized_rows)


__all__ = ["PDFStatementProcessor", "ProcessedTable", "StatementExtraction", "merge_tables"]
